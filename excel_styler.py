"""
COMAC Excel 样式引擎 - 国企汇报材料标准格式

遵循规范：
- 字体: 微软雅黑 (Microsoft YaHei), 统一字号
- 配色: 庄重忌花哨, 浅色底色区分层次
- 数据行: 纯白底色 (#FFFFFF)
- 机型编码: 浅绿底色 (#E8F5E9) 
- 重要列: 左置 (问题描述、根因分析、建议、团队)
- 次要列: 右置 (负责人、日期)
- 表头: 深色背景 (深蓝 #1A3A5C), 白色字体
"""
from typing import List, Dict, Tuple, Optional
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from pathlib import Path


# =============================================================================
# 样式常量 - COMAC 国企标准
# =============================================================================

# 字体
FONT_FAMILY = "Microsoft YaHei"  # 微软雅黑
FONT_HEADER = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
FONT_TITLE = Font(name=FONT_FAMILY, size=14, bold=True, color="1A3A5C")
FONT_SUBTITLE = Font(name=FONT_FAMILY, size=10, bold=True, color="333333")
FONT_BODY = Font(name=FONT_FAMILY, size=10, color="333333")
FONT_BOLD = Font(name=FONT_FAMILY, size=10, bold=True, color="333333")
FONT_SMALL = Font(name=FONT_FAMILY, size=9, color="666666")

# 填充色
FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_HEADER = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color="EBF0F7", end_color="EBF0F7", fill_type="solid")
FILL_LIGHT_GREEN = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FILL_LIGHT_YELLOW = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
FILL_LIGHT_GRAY = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
FILL_HIGHLIGHT = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

# 机型编码底色
FILL_C909 = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")      # 浅绿
FILL_C919 = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")      # 浅蓝
FILL_SIMILAR = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")   # 浅橙

# 对齐
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
HEADER_BORDER = Border(
    left=Side(style="thin", color="0A2A4A"),
    right=Side(style="thin", color="0A2A4A"),
    top=Side(style="thin", color="0A2A4A"),
    bottom=Side(style="medium", color="0A2A4A"),
)


@dataclass
class ColumnDef:
    """列定义"""
    name: str
    key: str
    width: int = 15
    importance: str = "normal"  # "high" 重要列左置 / "normal" / "low" 次要列右置
    wrap: bool = True


@dataclass
class SheetConfig:
    """工作表配置"""
    name: str
    columns: List[ColumnDef]
    title: str = ""
    subtitle: str = ""


class COMACExcelStyler:
    """
    COMAC 国企标准 Excel 样式引擎

    使用方式:
        styler = COMACExcelStyler()
        wb = styler.create_workbook("F项目经验教训总结")
        
        # 添加 C909 数据
        styler.add_sheet(wb, SheetConfig(
            name="C909",
            title="C909 动力装置系统经验教训",
            columns=[...]
        ), data_rows)
        
        styler.save(wb, "output.xlsx")
    """

    # 机型填充映射
    AIRCRAFT_FILLS = {
        "C909": FILL_C909,
        "C919": FILL_C919,
        "相似机型": FILL_SIMILAR,
        "c909": FILL_C909,
        "c919": FILL_C919,
    }

    def __init__(self):
        pass

    def create_workbook(self, report_title: str = "") -> Workbook:
        """创建标准化工作簿"""
        wb = Workbook()
        if report_title:
            ws = wb.active
            ws.title = "封面"
            self._create_cover(ws, report_title)
        return wb

    def _create_cover(self, ws, title: str):
        """创建封面"""
        ws.merge_cells("A1:G1")
        ws["A1"] = "中国商飞 COMAC"
        ws["A1"].font = Font(name=FONT_FAMILY, size=16, bold=True, color="1A3A5C")
        ws["A1"].alignment = ALIGN_CENTER

        ws.merge_cells("A2:G2")
        ws["A2"] = title
        ws["A2"].font = FONT_TITLE
        ws["A2"].alignment = ALIGN_CENTER

        ws.merge_cells("A4:G4")
        ws["A4"] = "— 内部资料 注意保密 —"
        ws["A4"].font = Font(name=FONT_FAMILY, size=9, color="999999")
        ws["A4"].alignment = ALIGN_CENTER

        ws.column_dimensions['A'].width = 15
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 12

    def add_sheet(
        self,
        wb: Workbook,
        config: SheetConfig,
        data_rows: List[Dict],
        aircraft_type: str = "",
    ) -> None:
        """
        添加标准化数据工作表

        Args:
            wb: 工作簿对象
            config: 工作表配置
            data_rows: 数据行列表
            aircraft_type: 机型标识 (用于底色编码)
        """
        ws = wb.create_sheet(title=config.name)

        # 按重要性排序: 重要列在左, 次要列在右
        sorted_cols = sorted(
            config.columns,
            key=lambda c: {"high": 0, "normal": 1, "low": 2}.get(c.importance, 1)
        )
        high_cols = [c for c in sorted_cols if c.importance == "high"]
        mid_cols = [c for c in sorted_cols if c.importance == "normal"]
        low_cols = [c for c in sorted_cols if c.importance == "low"]
        ordered_cols = high_cols + mid_cols + low_cols

        # 标题行
        if config.title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ordered_cols))
            ws["A1"] = config.title
            ws["A1"].font = FONT_TITLE
            ws["A1"].alignment = ALIGN_CENTER

        if config.subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ordered_cols))
            ws["A2"] = config.subtitle
            ws["A2"].font = FONT_SUBTITLE
            ws["A2"].alignment = ALIGN_CENTER

        # 表头行
        header_row = 3 if config.title else 1
        for col_idx, col_def in enumerate(ordered_cols, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_def.name)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = HEADER_BORDER

        # 列宽
        for col_idx, col_def in enumerate(ordered_cols, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_def.width

        # 获取机型填充色
        aircraft_fill = self.AIRCRAFT_FILLS.get(aircraft_type, FILL_WHITE)

        # 数据行
        data_start_row = header_row + 1
        for row_idx, row_data in enumerate(data_rows):
            excel_row = data_start_row + row_idx
            is_even = row_idx % 2 == 0

            for col_idx, col_def in enumerate(ordered_cols, start=1):
                value = row_data.get(col_def.key, "")
                cell = ws.cell(row=excel_row, column=col_idx, value=value)

                # 纯白底色数据行 (统一使用白色, 不用交替行色)
                cell.fill = FILL_WHITE
                cell.font = FONT_BODY
                cell.border = THIN_BORDER

                # 对齐: 重要列左对齐, 次要列居中
                if col_def.importance == "high":
                    cell.alignment = ALIGN_LEFT
                elif col_def.importance == "low":
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.alignment = ALIGN_LEFT if col_def.wrap else ALIGN_CENTER

            # 行高
            ws.row_dimensions[excel_row].height = 28

        # 汇总行 (浅蓝底色)
        summary_row = data_start_row + len(data_rows) + 1
        ws.cell(row=summary_row, column=1, value="合计").font = FONT_BOLD
        ws.cell(row=summary_row, column=1).fill = FILL_LIGHT_BLUE
        ws.cell(row=summary_row, column=1).alignment = ALIGN_CENTER
        for col_idx in range(2, len(ordered_cols) + 1):
            ws.cell(row=summary_row, column=col_idx).fill = FILL_LIGHT_BLUE
            ws.cell(row=summary_row, column=col_idx).border = THIN_BORDER

        # 冻结窗格
        ws.freeze_panes = ws.cell(row=data_start_row, column=1)

        # 打印设置
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A3

    def add_category_section(
        self,
        ws,
        start_row: int,
        category_name: str,
        category_fill: PatternFill,
        num_cols: int,
    ) -> int:
        """
        添加分类区域标题行 (浅色底色区分层次)

        Returns:
            数据起始行号
        """
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row, end_column=num_cols
        )
        cell = ws.cell(row=start_row, column=1, value=category_name)
        cell.font = FONT_BOLD
        cell.fill = category_fill
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER

        for col_idx in range(2, num_cols + 1):
            ws.cell(row=start_row, column=col_idx).fill = category_fill
            ws.cell(row=start_row, column=col_idx).border = THIN_BORDER

        return start_row + 1

    def apply_filter(self, ws, data_range_start: int, num_cols: int, last_row: int):
        """添加自动筛选"""
        ws.auto_filter.ref = f"A{data_range_start}:{get_column_letter(num_cols)}{last_row}"

    def save(self, wb: Workbook, filepath: str) -> str:
        """保存工作簿"""
        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return str(path)


# =============================================================================
# 预定义模板 - 航空发动机经验教训
# =============================================================================

def create_engine_lessons_template() -> Tuple[Workbook, COMACExcelStyler, List[SheetConfig]]:
    """
    创建航空发动机经验教训标准模板（完整9组合版）

    涵盖：C909 / C919 / 相似机型 × 发动机本体 / 短舱 / 飞发集成

    Returns:
        (Workbook, Styler, configs)
    """
    styler = COMACExcelStyler()
    wb = styler.create_workbook("F项目动力装置系统经验教训总结")

    # 统一的列定义（所有 sheet 共用）
    COLUMNS = [
        ColumnDef("序号", "seq", width=6, importance="low"),
        ColumnDef("问题描述", "problem", width=40, importance="high"),
        ColumnDef("根因分析", "root_cause", width=35, importance="high"),
        ColumnDef("改进建议", "suggestion", width=30, importance="high"),
        ColumnDef("责任团队", "team", width=15, importance="high"),
        ColumnDef("机型", "aircraft", width=8, importance="normal"),
        ColumnDef("产品类别", "category", width=12, importance="normal"),
        ColumnDef("严重程度", "severity", width=10, importance="normal"),
        ColumnDef("负责人", "owner", width=10, importance="low"),
        ColumnDef("完成日期", "date", width=12, importance="low"),
        ColumnDef("状态", "status", width=8, importance="low"),
    ]

    # 9 个组合：3 机型 × 3 产品类别
    AIRCRAFTS = ["C909", "C919", "相似机型"]
    CATEGORIES = ["发动机本体", "短舱", "飞发集成"]

    configs = []
    for ac in AIRCRAFTS:
        for cat in CATEGORIES:
            config = SheetConfig(
                name=f"{ac}-{cat}",
                title=f"{ac} 动力装置系统经验教训 — {cat}",
                subtitle="编制单位：动力装置部 | 密级：内部",
                columns=COLUMNS,
            )
            configs.append(config)

    return wb, styler, configs
